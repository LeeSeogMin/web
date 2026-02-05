#!/usr/bin/env python3
"""
contents.md를 강의계획서(web.xlsx) 형식으로 변환하는 스크립트
"""

import openpyxl
import re

def parse_contents():
    """contents.md 파일을 파싱하여 Chapter별 정보 추출"""
    with open('contents.md', 'r', encoding='utf-8') as f:
        content = f.read()

    chapters = []
    current_chapter = None
    current_part = None

    for line in content.split('\n'):
        # Part 추출
        if line.startswith('## Part'):
            current_part = line.replace('## ', '').strip()

        # Chapter 추출
        elif line.startswith('### Chapter'):
            if current_chapter:
                chapters.append(current_chapter)

            # "### Chapter 1. 웹이 동작하는 방식" 형식 파싱
            match = re.match(r'### Chapter (\d+)\. (.+)', line)
            if match:
                chapter_num = int(match.group(1))
                chapter_title = match.group(2).strip()
                current_chapter = {
                    'number': chapter_num,
                    'title': chapter_title,
                    'part': current_part,
                    'sections': []
                }

        # 절(section) 추출 - "1.1 클라이언트-서버 아키텍처" 형식
        elif current_chapter and re.match(r'^\d+\.\d+', line):
            section = line.strip()
            current_chapter['sections'].append(section)

    # 마지막 chapter 추가
    if current_chapter:
        chapters.append(current_chapter)

    return chapters

def get_josa(word, josa_type='eul'):
    """한글 단어에 맞는 조사 반환"""
    if not word:
        return ''

    # 마지막 글자의 유니코드 값
    last_char = word[-1]

    # 한글이 아닌 경우
    if not ('가' <= last_char <= '힣'):
        # 영문이나 숫자는 받침 없음으로 처리
        has_jongsung = False
    else:
        # 한글 유니코드: (초성 * 588) + (중성 * 28) + 종성 + 0xAC00
        code = ord(last_char) - 0xAC00
        jongsung = code % 28
        has_jongsung = jongsung != 0

    if josa_type == 'eul':  # 을/를
        return '을' if has_jongsung else '를'
    elif josa_type == 'i':  # 이/가
        return '이' if has_jongsung else '가'
    elif josa_type == 'euro':  # 으로/로
        return '으로' if has_jongsung else '로'
    else:
        return ''

def generate_learning_objectives(chapter):
    """Chapter 정보를 기반으로 학습목표 생성"""
    title = chapter['title']
    sections = chapter['sections']

    # Chapter 제목 기반 학습목표
    objectives = []

    # 주요 개념 파악
    josa = get_josa(title, 'eul')
    if '기초' in title or '기본' in title or '시작하기' in title:
        objectives.append(f"{title}의 핵심 개념을 이해한다")
    elif '심화' in title or '고급' in title:
        objectives.append(f"{title} 기법을 활용할 수 있다")
    elif '실전' in title or '프로젝트' in title:
        objectives.append(f"{title}{josa} 통해 실무 역량을 습득한다")
    elif '설계' in title:
        objectives.append(f"{title}의 원리와 방법을 학습한다")
    elif '구현' in title or '검증' in title:
        objectives.append(f"{title} 과정을 이해하고 적용할 수 있다")
    else:
        objectives.append(f"{title}{josa} 이해하고 활용할 수 있다")

    # 주요 섹션 기반 학습목표 추가 (최대 3-4개)
    key_topics = []
    for section in sections[:5]:  # 처음 5개 섹션만 사용
        # "1.1 클라이언트-서버 아키텍처" → "클라이언트-서버 아키텍처"
        topic = re.sub(r'^\d+\.\d+\s+', '', section)
        # 하위 섹션 제외
        if not topic.startswith('  -'):
            key_topics.append(topic)

    if key_topics:
        if len(key_topics) <= 2:
            for topic in key_topics:
                topic_josa = get_josa(topic, 'eul')
                objectives.append(f"{topic}{topic_josa} 설명할 수 있다")
        else:
            # 여러 주제를 하나로 묶기
            objectives.append(f"{', '.join(key_topics[:3])} 등의 개념을 학습한다")

    # 실습이 있는 경우
    practice_sections = [s for s in sections if '실습' in s]
    if practice_sections:
        objectives.append("실습을 통해 학습 내용을 실제로 구현할 수 있다")

    return '\n'.join(f"• {obj}" for obj in objectives[:4])  # 최대 4개

def generate_content_details(chapter):
    """Chapter의 세부 학습 내용 생성 (강의계획서용)"""
    sections = chapter['sections']

    # 주요 섹션만 추출 (하위 항목 제외)
    main_sections = []
    for section in sections:
        # 공백으로 시작하지 않는 섹션 (주요 섹션)
        if not section.startswith('  '):
            # 번호 제거하고 내용만 추출
            topic = re.sub(r'^\d+\.\d+\s+', '', section)
            main_sections.append(topic)

    # 강의계획서는 간결하게 - 최대 6개 정도만
    if len(main_sections) > 6:
        content = '\n'.join(f"- {s}" for s in main_sections[:5])
        content += f"\n- 외 {len(main_sections) - 5}개 주제"
    else:
        content = '\n'.join(f"- {s}" for s in main_sections)

    return content

def update_excel(chapters):
    """Excel 파일 업데이트"""
    from openpyxl.styles import Alignment

    wb = openpyxl.load_workbook('web.xlsx')
    ws = wb.active

    # 각 Chapter를 주차에 매핑
    for chapter in chapters:
        week_num = chapter['number']
        row_num = week_num + 1  # 1행은 헤더, 2행부터 데이터

        # 주차 (이미 있을 수 있음)
        week_cell = ws.cell(row=row_num, column=1)
        week_cell.value = f"{week_num:02d}"
        week_cell.number_format = '@'  # 텍스트 형식
        week_cell.alignment = Alignment(vertical='top', wrap_text=True)

        # 학습목표
        learning_objectives = generate_learning_objectives(chapter)
        obj_cell = ws.cell(row=row_num, column=2)
        obj_cell.value = learning_objectives
        obj_cell.number_format = '@'  # 텍스트 형식
        obj_cell.alignment = Alignment(vertical='top', wrap_text=True)

        # 교재명 및 페이지 정보 / 과제 및 평가마감 정보
        content_details = generate_content_details(chapter)
        content_cell = ws.cell(row=row_num, column=3)
        content_cell.value = content_details
        content_cell.number_format = '@'  # 텍스트 형식
        content_cell.alignment = Alignment(vertical='top', wrap_text=True)

    # 저장
    wb.save('web.xlsx')
    print(f"✅ 총 {len(chapters)}개 Chapter를 {len(chapters)}개 주차로 변환 완료!")
    print(f"📝 web.xlsx 파일이 업데이트되었습니다.")

def main():
    print("🔄 contents.md 파싱 중...")
    chapters = parse_contents()

    print(f"📚 총 {len(chapters)}개 Chapter 발견")
    for ch in chapters[:3]:  # 처음 3개만 미리보기
        print(f"  - Chapter {ch['number']}: {ch['title']} ({len(ch['sections'])}개 섹션)")
    print("  ...")

    print("\n📝 Excel 파일 업데이트 중...")
    update_excel(chapters)

if __name__ == '__main__':
    main()
